"""
The Closers - Business Finder
Searches Google Places, scores each website with Claude, and saves ONLY
true leads (no website, broken site, or missing contact/booking system).

Outputs two clean spreadsheets:
  leads_email.xlsx  — businesses where we found an email (pitch by email)
  leads_call.xlsx   — businesses with no email (pitch by phone)

Usage:
    python3 find_businesses.py "dental offices in Atlanta GA" --max 30
"""

import argparse
import os
import re
import time
import requests
import pandas as pd
from urllib.parse import urljoin, urlparse
from bs4 import BeautifulSoup
import anthropic

# ── API Keys ───────────────────────────────────────────────────────────────────
GOOGLE_API_KEY = "AIzaSyA-BrHr1cTOjB9-qyPNXb0CEw5BD7eG1g8"
# ──────────────────────────────────────────────────────────────────────────────

EMAIL_LEADS_FILE = "leads_email.xlsx"
CALL_LEADS_FILE  = "leads_call.xlsx"

PLACES_TEXT_SEARCH_URL = "https://places.googleapis.com/v1/places:searchText"
PLACES_DETAILS_URL     = "https://places.googleapis.com/v1/places/{place_id}"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    )
}

EMAIL_RE = re.compile(r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}")


# ── Google Places ──────────────────────────────────────────────────────────────

def search_places(query: str, max_results: int) -> list[dict]:
    places = []
    next_page_token = None
    while len(places) < max_results:
        payload = {"textQuery": query, "pageSize": min(20, max_results - len(places))}
        if next_page_token:
            payload["pageToken"] = next_page_token
        headers = {
            "Content-Type": "application/json",
            "X-Goog-Api-Key": GOOGLE_API_KEY,
            "X-Goog-FieldMask": "places.id,places.displayName,places.formattedAddress,places.rating,places.userRatingCount,nextPageToken",
        }
        response = requests.post(PLACES_TEXT_SEARCH_URL, json=payload, headers=headers, timeout=10)
        data = response.json()
        if "error" in data:
            print(f"\nERROR from Google Places: {data['error'].get('message', data['error'])}")
            return []
        places.extend(data.get("places", []))
        next_page_token = data.get("nextPageToken")
        if not next_page_token or len(places) >= max_results:
            break
        time.sleep(2)
    return places[:max_results]


def get_place_details(place_id: str) -> dict:
    url = PLACES_DETAILS_URL.format(place_id=place_id)
    headers = {"X-Goog-Api-Key": GOOGLE_API_KEY, "X-Goog-FieldMask": "nationalPhoneNumber,websiteUri"}
    response = requests.get(url, headers=headers, timeout=10)
    result = response.json()
    return {
        "phone":   result.get("nationalPhoneNumber", ""),
        "website": result.get("websiteUri", ""),
    }


# ── Website scraper ────────────────────────────────────────────────────────────

def scrape_website(url: str) -> dict:
    """Fetch a website and return its title, meta description, and body text."""
    empty = {"title": "", "description": "", "body_text": "", "status": "no_website"}
    if not url or not url.strip():
        return empty

    url = url.strip()
    if not url.startswith("http"):
        url = "https://" + url

    try:
        r = requests.get(url, headers=HEADERS, timeout=10)
        if r.status_code == 403:
            return {**empty, "status": "blocked", "body_text": "Site returned 403 Forbidden."}
        if r.status_code != 200:
            return {**empty, "status": f"error_{r.status_code}", "body_text": f"HTTP {r.status_code}"}
        soup = BeautifulSoup(r.text, "html.parser")
        title = soup.title.string.strip() if soup.title and soup.title.string else ""
        meta  = soup.find("meta", attrs={"name": "description"})
        desc  = meta.get("content", "").strip() if meta else ""
        for tag in soup(["script", "style", "nav", "footer", "header"]):
            tag.decompose()
        body = " ".join(soup.get_text(separator=" ").split())[:800]
        return {"title": title[:120], "description": desc[:200], "body_text": body, "status": "ok"}
    except requests.exceptions.SSLError:
        return {**empty, "status": "ssl_error", "body_text": "SSL certificate error."}
    except Exception as e:
        return {**empty, "status": "error", "body_text": str(e)[:100]}


# ── Claude website scorer ──────────────────────────────────────────────────────

def score_website(client: anthropic.Anthropic, business_name: str, business_type: str,
                  website: str, site: dict) -> dict:
    """
    Ask Claude to judge if this website is a lead worth pitching.
    Returns: {"is_lead": bool, "reason": str, "issue_summary": str}
    """
    # No website at all = automatic lead
    if not website or site["status"] == "no_website":
        return {
            "is_lead": True,
            "reason": "no_website",
            "issue_summary": "No website found — missing web presence entirely.",
        }

    # Broken/blocked = lead
    if site["status"] in ("blocked", "ssl_error") or site["status"].startswith("error_"):
        return {
            "is_lead": True,
            "reason": "broken_site",
            "issue_summary": f"Website is broken or inaccessible ({site['status']}) — visitors can't reach it.",
        }

    # Ask Claude to evaluate the site content
    prompt = f"""You are evaluating whether a local business is a good sales lead for The Closers.
The Closers sells two things:
1. A new professional website (main product)
2. An AI phone receptionist that books appointments 24/7 (secondary)

Business name: {business_name}
Business type: {business_type}
Website URL: {website}

Website content scraped:
Title: {site['title']}
Meta description: {site['description']}
Page text excerpt: {site['body_text'][:500]}

Evaluate this website and answer with EXACTLY this format, nothing else:

IS_LEAD: yes or no
REASON: one of: no_contact_form, no_booking_system, low_quality_content, broken_site, good_website
ISSUE: one sentence describing the specific problem (or "Website looks good" if no issue)

Rules:
- IS_LEAD = yes if: no way to contact/book, site looks very sparse/empty, content doesn't match the business type, or site seems low quality for their industry
- IS_LEAD = no if: site has a contact form OR booking system AND reasonable content for their industry
- Be strict — only flag real problems, not minor things
- A simple but functional site with a contact form = no (not a lead)
- A site with NO way to reach the business = yes (lead)"""

    try:
        response = client.messages.create(
            model="claude-opus-4-8",
            max_tokens=150,
            messages=[{"role": "user", "content": prompt}],
        )
        text = response.content[0].text.strip()

        is_lead = "IS_LEAD: yes" in text.lower()
        reason  = "unknown"
        issue   = "No specific issue identified"

        for line in text.split("\n"):
            if line.startswith("REASON:"):
                reason = line.replace("REASON:", "").strip()
            if line.startswith("ISSUE:"):
                issue = line.replace("ISSUE:", "").strip()

        return {"is_lead": is_lead, "reason": reason, "issue_summary": issue}

    except Exception as e:
        # If Claude fails, fall back to basic check
        has_contact = any(kw in site["body_text"].lower() for kw in
                         ["contact", "book", "schedule", "appointment", "call us", "email us"])
        return {
            "is_lead": not has_contact,
            "reason": "fallback_check",
            "issue_summary": "No contact or booking method found." if not has_contact else "Has contact method.",
        }


# ── Email finder ───────────────────────────────────────────────────────────────

def find_email(website: str) -> str:
    if not website or not website.strip():
        return ""
    url = website.strip()
    if not url.startswith("http"):
        url = "https://" + url
    base = f"{urlparse(url).scheme}://{urlparse(url).netloc}"
    pages = [url, urljoin(base, "/contact"), urljoin(base, "/contact-us"), urljoin(base, "/about")]
    seen = set()
    for page in pages:
        if page in seen:
            continue
        seen.add(page)
        try:
            r = requests.get(page, headers=HEADERS, timeout=8)
            if r.status_code != 200:
                continue
            soup = BeautifulSoup(r.text, "html.parser")
            for a in soup.find_all("a", href=True):
                href = a["href"]
                if href.startswith("mailto:"):
                    email = href.replace("mailto:", "").split("?")[0].strip()
                    if email and "@" in email:
                        return email
            for email in EMAIL_RE.findall(r.text):
                if any(s in email.lower() for s in [
                    "example.", "sentry.", "wix.", "squarespace.", "wordpress.",
                    ".png", ".jpg", ".gif", ".svg", ".webp", ".css", ".js",
                    "noreply", "no-reply", "donotreply",
                ]):
                    continue
                if "." not in email.split("@")[-1]:
                    continue
                return email
        except Exception:
            continue
    return ""


# ── Pricing tier classification (rule-based, no AI revenue guessing) ───────────

TIER_1_KEYWORDS = [
    "cafe", "coffee", "bakery", "food truck", "deli", "sandwich", "ice cream",
    "juice bar", "smoothie", "taco", "convenience store", "laundromat",
    "boutique", "thrift", "flower shop", "florist", "donut", "bagel",
    "snow cone", "popcorn", "candy shop", "small retail",
]

TIER_2_KEYWORDS = [
    "auto repair", "mechanic", "auto shop", "salon", "barbershop", "barber",
    "gym", "fitness", "clinic", "spa", "tattoo", "nail salon", "nails",
    "pet groom", "dry clean", "tire shop", "car wash", "yoga", "pilates",
    "massage", "chiropract", "physical therapy", "veterinary", "vet clinic",
    "tutoring", "daycare", "cleaning service", "landscap", "plumb", "hvac",
    "electrician", "locksmith", "moving company",
]

TIER_3_KEYWORDS = [
    "med spa", "medspa", "medical spa", "dentist", "dental", "dermatology",
    "cosmetic", "plastic surgery", "orthodont", "realtor", "real estate",
    "financial advisor", "accountant", "cpa", "insurance agency",
    "wealth management", "consulting firm", "architect", "engineering firm",
    "private practice", "wellness center", "aesthetic", "concierge",
]

EXCLUDE_LAW_KEYWORDS = [
    "law firm", "law office", "attorney", "lawyer", "legal", "esq",
    "law group", "pllc law", "law pllc", "criminal defense", "injury law",
]

# Known large chains / franchises to exclude outright
CHAIN_KEYWORDS = [
    "the joint chiropractic", "massage envy", "european wax center",
    "great clips", "supercuts", "sport clips", "jiffy lube", "midas",
    "firestone", "pep boys", "meineke", "valvoline", "planet fitness",
    "anytime fitness", "orangetheory", "crunch fitness", "la fitness",
    "starbucks", "dunkin", "subway", "mcdonald's", "chipotle",
    "h&r block", "fastsigns", "ups store", "fedex office",
]


def should_exclude(name: str, business_type: str, rating_count) -> dict:
    """
    Decide if a business should be excluded entirely (law firms, large/chain businesses).
    Returns {"exclude": bool, "reason": str, "uncertain": bool}
    """
    name_lower = name.lower()
    type_lower = business_type.lower()

    # Exclude law firms outright
    if any(kw in name_lower or kw in type_lower for kw in EXCLUDE_LAW_KEYWORDS):
        return {"exclude": True, "reason": "Law firm — excluded by category", "uncertain": False}

    # Exclude known chains outright
    if any(kw in name_lower for kw in CHAIN_KEYWORDS):
        return {"exclude": True, "reason": "Recognized national chain — excluded", "uncertain": False}

    # Review count signals
    count = 0
    try:
        count = int(rating_count) if rating_count else 0
    except (ValueError, TypeError):
        count = 0

    if count >= 300:
        return {"exclude": True, "reason": f"{count} reviews — likely established/large business", "uncertain": False}

    if count >= 150:
        return {
            "exclude": False, "uncertain": True,
            "reason": f"Uncertain size — {count} reviews, verify before pitching",
        }

    return {"exclude": False, "reason": "", "uncertain": False}


def classify_tier(name: str, business_type: str) -> dict:
    """
    Rule-based pricing tier based on business category keywords.
    Returns {"tier": str, "setup_price": str, "retainer": str, "uncertain": bool}
    """
    name_lower = name.lower()
    type_lower = business_type.lower()
    combined   = f"{name_lower} {type_lower}"

    if any(kw in combined for kw in TIER_1_KEYWORDS):
        return {"tier": "Tier 1", "setup_price": "$150–250", "retainer": "$50–75/mo", "uncertain": False}

    if any(kw in combined for kw in TIER_3_KEYWORDS):
        return {"tier": "Tier 3", "setup_price": "$500–800+", "retainer": "$150–250/mo", "uncertain": False}

    if any(kw in combined for kw in TIER_2_KEYWORDS):
        return {"tier": "Tier 2", "setup_price": "$300–500", "retainer": "$100–150/mo", "uncertain": False}

    # No clear match — default to Tier 2, flag for manual review
    return {
        "tier": "Tier 2 (default)", "setup_price": "$300–500", "retainer": "$100–150/mo",
        "uncertain": True,
    }


# ── Recommended price within tier (rule-based, using review count as a maturity signal) ─

TIER_PRICE_BRACKETS = {
    "Tier 1": [(10, 150, 50), (30, 200, 60), (999, 250, 75)],
    "Tier 2": [(15, 300, 100), (50, 400, 125), (999, 500, 150)],
    "Tier 3": [(20, 500, 150), (60, 650, 200), (999, 800, 250)],
}


def estimate_price(tier: str, rating_count) -> dict:
    """
    Recommend a SPECIFIC dollar amount within the tier's range, biased toward
    the low end for newer/smaller businesses (few reviews = likely can't
    afford much yet) and the high end for more established ones.
    This is NOT a revenue guess — it only uses review count as a rough
    maturity signal, the same logic a person would use eyeballing a listing.
    """
    base_tier = tier.replace(" (default)", "")
    brackets  = TIER_PRICE_BRACKETS.get(base_tier, TIER_PRICE_BRACKETS["Tier 2"])

    count = 0
    try:
        count = int(rating_count) if rating_count else 0
    except (ValueError, TypeError):
        count = 0

    for max_count, setup, retainer in brackets:
        if count <= max_count:
            if count <= brackets[0][0]:
                maturity = "Looks new/small (few reviews) — recommend the low end to close the sale"
            elif count >= brackets[-2][0]:
                maturity = "Looks more established — can likely support the higher end"
            else:
                maturity = "Moderate size — mid-range pricing"
            return {
                "recommended_setup":    f"${setup}",
                "recommended_retainer": f"${retainer}/mo",
                "maturity_note":        maturity,
            }

    # Fallback (shouldn't hit since last bracket is 999)
    setup, retainer = brackets[-1][1], brackets[-1][2]
    return {
        "recommended_setup":    f"${setup}",
        "recommended_retainer": f"${retainer}/mo",
        "maturity_note":        "Moderate size — mid-range pricing",
    }


# ── Multi-location / duplicate detection ───────────────────────────────────────

def normalize_name(name: str) -> str:
    """Lowercase, strip punctuation, collapse whitespace."""
    name = name.lower()
    name = re.sub(r"[^a-z0-9\s]", "", name)
    return re.sub(r"\s+", " ", name).strip()


def is_likely_same_business(name_a: str, name_b: str) -> bool:
    """
    Catch near-duplicate listings Google Places sometimes returns for the
    same business under slightly different name formatting, e.g.
    'Holston & Huntley' vs 'Holston & Huntley Trial Attorneys'.
    """
    a, b = normalize_name(name_a), normalize_name(name_b)
    if not a or not b:
        return False
    shorter, longer = (a, b) if len(a) <= len(b) else (b, a)
    if len(shorter) < 5:
        return False
    return shorter in longer


def detect_multi_location(places: list[dict]) -> set:
    """
    Return the set of place indices that appear to be the same business
    listed multiple times (exact duplicate OR multi-location chain).
    """
    names = [p.get("displayName", {}).get("text", "") for p in places]
    assigned = [False] * len(names)
    groups   = []

    for i in range(len(names)):
        if assigned[i]:
            continue
        group = [i]
        assigned[i] = True
        for j in range(i + 1, len(names)):
            if assigned[j]:
                continue
            if is_likely_same_business(names[i], names[j]):
                group.append(j)
                assigned[j] = True
        groups.append(group)

    duplicate_indices = set()
    for group in groups:
        if len(group) > 1:
            duplicate_indices.update(group)
    return duplicate_indices


# ── Spreadsheet writer ─────────────────────────────────────────────────────────

def append_to_excel(file: str, new_rows: list[dict], dedup_cols: list[str]):
    """Append rows to an Excel file, deduplicating on the given columns."""
    new_df = pd.DataFrame(new_rows)
    try:
        existing = pd.read_excel(file)
        combined = pd.concat([existing, new_df], ignore_index=True)
        combined = combined.drop_duplicates(subset=dedup_cols, keep="first")
    except FileNotFoundError:
        combined = new_df

    with pd.ExcelWriter(file, engine="openpyxl") as writer:
        combined.to_excel(writer, index=False, sheet_name="Leads")
        ws = writer.sheets["Leads"]
        # Style: freeze top row, bold headers
        from openpyxl.styles import Font, PatternFill, Alignment
        for cell in ws[1]:
            cell.font      = Font(bold=True, color="FFFFFF")
            cell.fill      = PatternFill("solid", fgColor="2F4F4F")
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.freeze_panes = "A2"
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col) + 4
            ws.column_dimensions[col[0].column_letter].width = min(max_len, 55)
        # Wrap text in all data rows
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[1].height = 30

    return len(combined)


# ── Main ───────────────────────────────────────────────────────────────────────

def run(query: str, max_results: int):
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        print("ERROR: ANTHROPIC_API_KEY not set. Run: export ANTHROPIC_API_KEY='your-key'")
        return

    client = anthropic.Anthropic(api_key=api_key)

    print(f"\nSearching Google Places for: \"{query}\"")
    print(f"Max results: {max_results}\n")

    places = search_places(query, max_results)
    if not places:
        print("No results found.")
        return

    # Infer business type from query for better scoring
    business_type = query.split(" in ")[0].strip() if " in " in query else query

    # Detect businesses that appear more than once (exact dupes or multi-location chains)
    multi_location_indices = detect_multi_location(places)

    print(f"Found {len(places)} businesses. Scoring each one...\n")

    email_leads = []
    call_leads  = []
    skipped     = 0
    excluded    = 0

    for i, place in enumerate(places, start=1):
        name         = place.get("displayName", {}).get("text", "")
        address      = place.get("formattedAddress", "")
        rating_count = place.get("userRatingCount", 0)

        print(f"[{i}/{len(places)}] {name}")

        # Skip duplicate/multi-location listings entirely
        if (i - 1) in multi_location_indices:
            print(f"         ⛔ Excluded — appears multiple times (duplicate listing or multi-location chain)")
            excluded += 1
            continue

        # Check exclusion (law firms, large/chain businesses) BEFORE scoring website
        exclusion = should_exclude(name, business_type, rating_count)
        if exclusion["exclude"]:
            print(f"         ⛔ Excluded — {exclusion['reason']}")
            excluded += 1
            continue

        details = get_place_details(place["id"])
        phone   = details["phone"]
        website = details["website"]

        # Scrape the website
        site = scrape_website(website)

        # Ask Claude if this is a lead
        print(f"         Scoring website...", end=" ", flush=True)
        score = score_website(client, name, business_type, website, site)

        if not score["is_lead"]:
            print(f"✓ Good website — skipping")
            skipped += 1
            time.sleep(0.3)
            continue

        print(f"✗ LEAD — {score['issue_summary'][:60]}")

        # Classify pricing tier (rule-based on category keywords)
        tier_info  = classify_tier(name, business_type)
        price_info = estimate_price(tier_info["tier"], rating_count)

        # Build notes — combine size-uncertainty flag, tier-uncertainty flag, and maturity note
        notes_parts = [price_info["maturity_note"]]
        if exclusion["uncertain"]:
            notes_parts.append(exclusion["reason"])
        if tier_info["uncertain"]:
            notes_parts.append("No clear category match for tier — defaulted to Tier 2, please verify.")
        notes = " | ".join(notes_parts)

        # Find email
        email = find_email(website) if website else ""

        if email:
            email_leads.append({
                "Business Name":      name,
                "Phone":              phone,
                "Website":            website or "(none)",
                "Email":              email,
                "Suggested Tier":     tier_info["tier"],
                "Recommended Price":  price_info["recommended_setup"],
                "Recommended Retainer": price_info["recommended_retainer"],
                "Price Range":        f"{tier_info['setup_price']} setup / {tier_info['retainer']} retainer",
                "Issue":              score["issue_summary"],
                "My Notes":           notes,
            })
        else:
            call_leads.append({
                "Business Name":      name,
                "Phone":              phone,
                "Website":            website or "(none)",
                "Suggested Tier":     tier_info["tier"],
                "Recommended Price":  price_info["recommended_setup"],
                "Recommended Retainer": price_info["recommended_retainer"],
                "Price Range":        f"{tier_info['setup_price']} setup / {tier_info['retainer']} retainer",
                "Issue":              score["issue_summary"],
                "My Notes":           notes,
                "Call Script":        "",  # filled by generate_pitches.py
            })

        time.sleep(0.5)

    # Save both spreadsheets — dedup on Email/Phone alone since that's the
    # strongest unique signal (catches the same business showing up under a
    # slightly different name in a later search)
    if email_leads:
        total = append_to_excel(EMAIL_LEADS_FILE, email_leads, ["Email"])
        print(f"\n✉  {len(email_leads)} email leads saved to '{EMAIL_LEADS_FILE}' ({total} total)")

    if call_leads:
        total = append_to_excel(CALL_LEADS_FILE, call_leads, ["Phone"])
        print(f"📞  {len(call_leads)} call leads saved to '{CALL_LEADS_FILE}' ({total} total)")

    print(f"\n{skipped} businesses skipped (good websites).")
    print(f"{excluded} businesses excluded (law firms / large chains).")
    print(f"Total leads found: {len(email_leads) + len(call_leads)}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("query", help='e.g. "dental offices in Atlanta GA"')
    parser.add_argument("--max", type=int, default=20)
    args = parser.parse_args()
    run(args.query, args.max)
