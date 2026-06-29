"""
The Closers - Pitch Generator
Reads leads_email.xlsx and leads_call.xlsx and generates:
  - Personalized pitch emails for email leads
  - Cold call scripts for call leads

Usage:
    python3 generate_pitches.py
"""

import os
import time
import requests
import pandas as pd
from bs4 import BeautifulSoup
import anthropic

EMAIL_LEADS_FILE = "leads_email.xlsx"
CALL_LEADS_FILE  = "leads_call.xlsx"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    )
}


def scrape_website_summary(url: str) -> str:
    """Return a short text snapshot of a website for context."""
    if not url or url == "(none)":
        return ""
    if not url.startswith("http"):
        url = "https://" + url
    try:
        r = requests.get(url, headers=HEADERS, timeout=8)
        if r.status_code != 200:
            return ""
        soup = BeautifulSoup(r.text, "html.parser")
        for tag in soup(["script", "style", "nav", "footer", "header"]):
            tag.decompose()
        return " ".join(soup.get_text(separator=" ").split())[:500]
    except Exception:
        return ""


def generate_email(client: anthropic.Anthropic, name: str, website: str,
                   issue: str, body_text: str) -> str:
    prompt = f"""You write cold outreach emails for "The Closers," a service that builds professional websites and sets up AI phone receptionists for local businesses.

Business: {name}
Website: {website}
Problem found: {issue}
Website content: {body_text[:400] if body_text else "(no website or inaccessible)"}

Write a SHORT cold email (4-6 sentences) that:
- Opens referencing something real and specific about their business
- Names the specific problem you found in one plain sentence
- Explains what The Closers can fix and why it matters for their type of business
- Ends with a soft CTA like "Worth a quick 10-min call?"
- Sounds like a real person — conversational, no buzzwords, no emojis
- Does NOT include subject line, sign-off, or placeholder brackets

Just the email body. Nothing else."""

    r = client.messages.create(
        model="claude-opus-4-8",
        max_tokens=400,
        messages=[{"role": "user", "content": prompt}],
    )
    return r.content[0].text.strip()


def generate_call_script(client: anthropic.Anthropic, name: str, website: str,
                         issue: str) -> str:
    prompt = f"""You create cold call scripts for "The Closers," a service that builds professional websites and sets up AI phone receptionists for local businesses.

Business: {name}
Website: {website or "(none)"}
Problem found: {issue}

Write a cold call script in this format:

OPENING (1-2 sentences to get their attention — reference the specific issue)

KEY POINTS (3-4 bullet points on why this hurts their business and what The Closers fixes)

OBJECTION RESPONSES:
- "I already have a website" → (response)
- "I'm not interested" → (response)
- "I don't have the budget" → (response)
- "Send me an email" → (response)

CLOSE (1 sentence to end the call with a next step)

Keep it natural and conversational. Don't make it sound like a script being read."""

    r = client.messages.create(
        model="claude-opus-4-8",
        max_tokens=600,
        messages=[{"role": "user", "content": prompt}],
    )
    return r.content[0].text.strip()


def save_excel(df: pd.DataFrame, file: str):
    with pd.ExcelWriter(file, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Leads")
        ws = writer.sheets["Leads"]
        from openpyxl.styles import Font, PatternFill, Alignment
        for cell in ws[1]:
            cell.font      = Font(bold=True, color="FFFFFF")
            cell.fill      = PatternFill("solid", fgColor="2F4F4F")
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.freeze_panes = "A2"
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col) + 4
            ws.column_dimensions[col[0].column_letter].width = min(max_len, 60)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[1].height = 30


def process_email_leads(client: anthropic.Anthropic):
    try:
        df = pd.read_excel(EMAIL_LEADS_FILE)
    except FileNotFoundError:
        print("No leads_email.xlsx found — run find_businesses.py first.")
        return

    if "Draft Email" not in df.columns:
        df["Draft Email"] = ""
    if "Email Status" not in df.columns:
        df["Email Status"] = ""
    if "Sent At" not in df.columns:
        df["Sent At"] = ""

    pending = df[df["Draft Email"].apply(
        lambda x: pd.isna(x) or str(x).strip() == "" or str(x).startswith("[Error")
    )].index.tolist()

    if not pending:
        print("✉  All email leads already have drafts.")
        return

    print(f"✉  Writing emails for {len(pending)} leads...\n")

    for i, idx in enumerate(pending, start=1):
        row  = df.loc[idx]
        name = str(row.get("Business Name", "") or "")
        website = str(row.get("Website", "") or "")
        issue   = str(row.get("Issue", "") or "")

        print(f"  [{i}/{len(pending)}] {name}...", end=" ", flush=True)
        body_text = scrape_website_summary(website)

        try:
            email = generate_email(client, name, website, issue, body_text)
            df.at[idx, "Draft Email"] = email
            print("done")
        except Exception as e:
            df.at[idx, "Draft Email"] = f"[Error: {e}]"
            print(f"error: {e}")

        time.sleep(0.5)

    save_excel(df, EMAIL_LEADS_FILE)
    print(f"\n  Saved to '{EMAIL_LEADS_FILE}'")


def process_call_leads(client: anthropic.Anthropic):
    try:
        df = pd.read_excel(CALL_LEADS_FILE)
    except FileNotFoundError:
        print("No leads_call.xlsx found — run find_businesses.py first.")
        return

    if "Call Script" not in df.columns:
        df["Call Script"] = ""

    pending = df[df["Call Script"].apply(
        lambda x: pd.isna(x) or str(x).strip() == "" or str(x).startswith("[Error")
    )].index.tolist()

    if not pending:
        print("📞  All call leads already have scripts.")
        return

    print(f"📞  Writing call scripts for {len(pending)} leads...\n")

    for i, idx in enumerate(pending, start=1):
        row     = df.loc[idx]
        name    = str(row.get("Business Name", "") or "")
        website = str(row.get("Website", "") or "")
        issue   = str(row.get("Issue", "") or "")

        print(f"  [{i}/{len(pending)}] {name}...", end=" ", flush=True)

        try:
            script = generate_call_script(client, name, website, issue)
            df.at[idx, "Call Script"] = script
            print("done")
        except Exception as e:
            df.at[idx, "Call Script"] = f"[Error: {e}]"
            print(f"error: {e}")

        time.sleep(0.5)

    save_excel(df, CALL_LEADS_FILE)
    print(f"\n  Saved to '{CALL_LEADS_FILE}'")


def run():
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        print("ERROR: ANTHROPIC_API_KEY not set. Run: export ANTHROPIC_API_KEY='your-key'")
        return

    client = anthropic.Anthropic(api_key=api_key)
    print("Generating pitches...\n")
    process_email_leads(client)
    print()
    process_call_leads(client)
    print("\nDone! Review your leads before reaching out.")


if __name__ == "__main__":
    run()
