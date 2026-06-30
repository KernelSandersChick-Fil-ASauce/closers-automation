"""
The Closers - PDF Audit Generator
Generates a clean, one-page "Free Online Presence Audit" PDF for every
business flagged for calling (leads_call.xlsx). Ready to email the moment
a business says yes on the call.

Usage:
    python3 generate_audits.py
"""

import os
import re
import pandas as pd
import anthropic
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.enums import TA_LEFT, TA_CENTER

CALL_LEADS_FILE = "leads_call.xlsx"
AUDIT_FOLDER    = "audits"

# Brand colors
DARK    = colors.HexColor("#2F4F4F")
ACCENT  = colors.HexColor("#C75B39")
LIGHT   = colors.HexColor("#F4F1EA")


def sanitize_filename(name: str) -> str:
    """Turn a business name into a safe filename."""
    name = re.sub(r"[^\w\s-]", "", name).strip()
    name = re.sub(r"[-\s]+", "_", name)
    return name[:60]


def generate_audit_content(client: anthropic.Anthropic, name: str, website: str, issue: str) -> dict:
    """
    Ask Claude for 2-3 'What We Found' bullets and a 2-3 sentence
    'What This Is Costing You' paragraph, based on the detected issue.
    """
    prompt = f"""You write content for a one-page audit PDF that The Closers sends to local businesses as a free gift after a cold call.

Business name: {name}
Website: {website or "No website found"}
Issue detected by our automated check: {issue}

Write two things:

1. WHAT_WE_FOUND — 2-3 short bullet points (plain text, no dashes or bullet symbols, one per line) describing what's missing or broken. Be factual and specific, not salesy.

2. WHAT_ITS_COSTING — 2-3 plain-English sentences explaining what this gap actually costs the business in real terms (missed calls, lost customers, looking unprofessional, etc.) — tailored to what kind of business this is. No buzzwords, no exaggeration, just a clear honest explanation.

Respond in EXACTLY this format:
WHAT_WE_FOUND:
[bullet 1]
[bullet 2]
[bullet 3 if applicable]

WHAT_ITS_COSTING:
[paragraph text]"""

    response = client.messages.create(
        model="claude-opus-4-8",
        max_tokens=400,
        messages=[{"role": "user", "content": prompt}],
    )
    text = response.content[0].text.strip()

    bullets = []
    costing = ""
    section = None
    for line in text.split("\n"):
        line = line.strip()
        if line.startswith("WHAT_WE_FOUND:"):
            section = "bullets"
            continue
        if line.startswith("WHAT_ITS_COSTING:"):
            section = "costing"
            continue
        if not line:
            continue
        if section == "bullets":
            bullets.append(line.lstrip("-•* ").strip())
        elif section == "costing":
            costing += (" " if costing else "") + line

    if not bullets:
        bullets = [issue] if issue else ["No clear way for customers to reach or book this business online."]
    if not costing:
        costing = "Missing this is likely costing the business new customers every week."

    return {"bullets": bullets[:3], "costing": costing}


def build_pdf(filepath: str, business_name: str, website: str, content: dict):
    """Build the one-page audit PDF."""
    doc = SimpleDocTemplate(
        filepath, pagesize=letter,
        topMargin=0.6 * inch, bottomMargin=0.6 * inch,
        leftMargin=0.75 * inch, rightMargin=0.75 * inch,
    )

    styles = getSampleStyleSheet()

    header_style = ParagraphStyle(
        "HeaderStyle", parent=styles["Title"],
        fontSize=22, textColor=DARK, alignment=TA_LEFT, spaceAfter=4,
    )
    subheader_style = ParagraphStyle(
        "SubHeader", parent=styles["Normal"],
        fontSize=13, textColor=ACCENT, spaceAfter=18, fontName="Helvetica-Bold",
    )
    section_title_style = ParagraphStyle(
        "SectionTitle", parent=styles["Heading2"],
        fontSize=13, textColor=DARK, spaceBefore=18, spaceAfter=8,
    )
    body_style = ParagraphStyle(
        "Body", parent=styles["Normal"],
        fontSize=11, leading=16, textColor=colors.HexColor("#333333"),
    )
    bullet_style = ParagraphStyle(
        "Bullet", parent=body_style,
        leftIndent=16, spaceAfter=6,
    )
    footer_style = ParagraphStyle(
        "Footer", parent=styles["Normal"],
        fontSize=9, textColor=colors.HexColor("#888888"), alignment=TA_CENTER,
    )

    story = []

    # Header
    story.append(Paragraph("Your Free Online Presence Audit", header_style))
    story.append(Paragraph(business_name, subheader_style))

    # Divider line via table
    story.append(Table([[""]], colWidths=[6.5 * inch], rowHeights=[2],
                        style=TableStyle([("BACKGROUND", (0, 0), (-1, -1), ACCENT)])))
    story.append(Spacer(1, 16))

    # What We Found
    story.append(Paragraph("What We Found", section_title_style))
    for bullet in content["bullets"]:
        story.append(Paragraph(f"•  {bullet}", bullet_style))

    # What This Is Costing You
    story.append(Paragraph("What This Is Costing You", section_title_style))
    story.append(Paragraph(content["costing"], body_style))

    # CTA box
    story.append(Spacer(1, 24))
    cta_table = Table(
        [["We fix this. Simple, professional websites and AI phone receptionists\n"
          "for local businesses — built fast, priced fairly.\n\n"
          "Reply to this email or call us back to get started."]],
        colWidths=[6.5 * inch],
        style=TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), LIGHT),
            ("BOX", (0, 0), (-1, -1), 1, ACCENT),
            ("LEFTPADDING", (0, 0), (-1, -1), 14),
            ("RIGHTPADDING", (0, 0), (-1, -1), 14),
            ("TOPPADDING", (0, 0), (-1, -1), 12),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 12),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]),
    )
    story.append(cta_table)

    story.append(Spacer(1, 30))
    story.append(Paragraph("The Closers — hqsummitos@gmail.com", footer_style))

    doc.build(story)


def run():
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        print("ERROR: ANTHROPIC_API_KEY not set. Run: export ANTHROPIC_API_KEY='your-key'")
        return

    client = anthropic.Anthropic(api_key=api_key)

    try:
        df = pd.read_excel(CALL_LEADS_FILE)
    except FileNotFoundError:
        print(f"No '{CALL_LEADS_FILE}' found — run find_businesses.py first.")
        return

    if "Audit PDF" not in df.columns:
        df["Audit PDF"] = ""

    os.makedirs(AUDIT_FOLDER, exist_ok=True)

    pending = df[df["Audit PDF"].apply(
        lambda x: pd.isna(x) or str(x).strip() == "" or not os.path.exists(str(x))
    )].index.tolist()

    if not pending:
        print("All call leads already have audit PDFs ready.")
        return

    print(f"Generating audit PDFs for {len(pending)} businesses...\n")

    for i, idx in enumerate(pending, start=1):
        row     = df.loc[idx]
        name    = str(row.get("Business Name", "") or "Unknown Business")
        website = str(row.get("Website", "") or "")
        issue   = str(row.get("Issue", "") or "")

        print(f"[{i}/{len(pending)}] {name}...", end=" ", flush=True)

        try:
            content  = generate_audit_content(client, name, website, issue)
            filename = sanitize_filename(name) + ".pdf"
            filepath = os.path.join(AUDIT_FOLDER, filename)
            build_pdf(filepath, name, website, content)
            df.at[idx, "Audit PDF"] = filepath
            print("done")
        except Exception as e:
            print(f"error: {e}")

    # Save updated spreadsheet
    with pd.ExcelWriter(CALL_LEADS_FILE, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Leads")
        ws = writer.sheets["Leads"]
        from openpyxl.styles import Font, PatternFill, Alignment
        for cell in ws[1]:
            cell.font      = Font(bold=True, color="FFFFFF")
            cell.fill      = PatternFill("solid", fgColor="2F4F4F")
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.freeze_panes = "A2"
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col) + 4
            ws.column_dimensions[col[0].column_letter].width = min(max_len, 60)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[1].height = 30

    print(f"\nDone! Audit PDFs saved in '{AUDIT_FOLDER}/' folder.")
    print(f"Spreadsheet updated with file paths in the 'Audit PDF' column.")


if __name__ == "__main__":
    run()
